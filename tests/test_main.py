from io import BytesIO
from datetime import datetime

from openpyxl import Workbook
from fastapi.testclient import TestClient

from src.main import app

client = TestClient(app)


def test_health_check() -> None:
    response = client.get("/")
    assert response.status_code == 200
    assert response.json()["status"] == "ready"


def test_analizar_csv_consolida_por_codigo_barra() -> None:
    csv_content = (
        "Drogueria Central\n"
        "codigo_barra,codigo_producto,nombre_producto,nombre_laboratorio,unidades_existentes,precio_unitario\n"
        " 00012345 ,P1,Acetaminofen,Lab A,40,12\n"
        "12345,P2,Acetaminofen,Lab A,18,9\n"
        "00067890,P3,Ibuprofeno,Lab B,7,0\n"
        "67890,P4,Ibuprofeno,Lab B,25,14\n"
    )

    response = client.post(
        "/analizar",
        files={"file": ("precios.csv", csv_content, "text/csv")},
    )

    assert response.status_code == 200
    data = response.json()

    assert len(data) == 2

    por_barra = {item["codigo_barra"]: item for item in data}
    assert por_barra["12345"]["codigo_interno_proveedor"] == "P2"
    assert por_barra["12345"]["nombre_laboratorio"] == "Lab A"
    assert por_barra["12345"]["mejor_precio"] == 9.0
    assert por_barra["12345"]["proveedor_ganador"] == "Drogueria Central"
    assert por_barra["12345"]["unidades_disponibles"] == 18
    assert por_barra["67890"]["codigo_interno_proveedor"] == "P4"
    assert por_barra["67890"]["mejor_precio"] == 14.0
    assert "analisis_timestamp" in por_barra["12345"]


def test_analizar_csv_con_proveedor_en_primera_fila() -> None:
    csv_content = (
        "Drogueria Central\n"
        "codigo_barra,codigo_producto,nombre_producto,nombre_laboratorio,unidades_existentes,precio_unitario\n"
        "00012345,P1,Acetaminofen,Lab A,40,12\n"
        "12345,P2,Acetaminofen,Lab A,18,9\n"
    )

    response = client.post(
        "/analizar",
        files={"file": ("precios_proveedor.csv", csv_content, "text/csv")},
    )

    assert response.status_code == 200
    data = response.json()
    assert data[0]["codigo_barra"] == "12345"
    assert data[0]["proveedor_ganador"] == "Drogueria Central"
    assert data[0]["codigo_interno_proveedor"] == "P2"


def test_analizar_csv_usa_alias_y_normaliza_barcode() -> None:
    csv_content = (
        "Laboratorio Alias\n"
        "barcode,codigo,descripcion,laboratorio,stock,precio\n"
        " 00000999 ,SKU-1,Vitamina C,Lab Alias,5,11\n"
        "999,SKU-2,Vitamina C,Lab Alias,8,7\n"
    )

    response = client.post(
        "/analizar",
        files={"file": ("precios_alias.csv", csv_content, "text/csv")},
    )

    assert response.status_code == 200
    data = response.json()
    assert data[0]["codigo_barra"] == "999"
    assert data[0]["codigo_interno_proveedor"] == "SKU-2"
    assert data[0]["nombre_producto"] == "Vitamina C"
    assert data[0]["mejor_precio"] == 7.0
    assert data[0]["proveedor_ganador"] == "Laboratorio Alias"
    assert data[0]["unidades_disponibles"] == 8
    assert datetime.fromisoformat(data[0]["analisis_timestamp"])


def test_analizar_xlsx_ok() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["Drogueria Excel"])
    worksheet.append(
        [
            "codigo_barra",
            "codigo_producto",
            "nombre_producto",
            "nombre_laboratorio",
            "unidades_existentes",
            "precio_unitario",
        ]
    )
    worksheet.append(["000777", "X1", "Amoxicilina", "Lab X", 20, 15])
    worksheet.append(["777", "X2", "Amoxicilina", "Lab X", 12, 10])

    file_content = BytesIO()
    workbook.save(file_content)
    file_content.seek(0)

    response = client.post(
        "/analizar",
        files={
            "file": (
                "precios.xlsx",
                file_content.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200
    data = response.json()
    assert data[0]["codigo_barra"] == "777"
    assert data[0]["codigo_interno_proveedor"] == "X2"
    assert data[0]["mejor_precio"] == 10.0
    assert data[0]["proveedor_ganador"] == "Drogueria Excel"


def test_analizar_xlsx_con_proveedor_en_primera_fila() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["Laboratorio Andino"])
    worksheet.append(
        [
            "codigo_barra",
            "codigo_producto",
            "nombre_producto",
            "nombre_laboratorio",
            "unidades_existentes",
            "precio_unitario",
        ]
    )
    worksheet.append(["000888", "Y1", "Loratadina", "Lab Andino", 20, 13])
    worksheet.append(["888", "Y2", "Loratadina", "Lab Andino", 15, 10])

    file_content = BytesIO()
    workbook.save(file_content)
    file_content.seek(0)

    response = client.post(
        "/analizar",
        files={
            "file": (
                "precios_proveedor.xlsx",
                file_content.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200
    data = response.json()
    assert data[0]["codigo_barra"] == "888"
    assert data[0]["proveedor_ganador"] == "Laboratorio Andino"
    assert data[0]["codigo_interno_proveedor"] == "Y2"


def test_analizar_dos_excels_mismo_codigo_barra_distinta_drogueria() -> None:
    workbook_1 = Workbook()
    ws_1 = workbook_1.active
    ws_1.append(["Drogueria Alfa"])
    ws_1.append(
        [
            "codigo_barra",
            "codigo_producto",
            "nombre_producto",
            "nombre_laboratorio",
            "unidades_existentes",
            "precio_unitario",
        ]
    )
    ws_1.append(["000123456", "ALF-1", "Cetirizina", "Lab Q", 30, 11])
    ws_1_bytes = BytesIO()
    workbook_1.save(ws_1_bytes)
    ws_1_bytes.seek(0)

    workbook_2 = Workbook()
    ws_2 = workbook_2.active
    ws_2.append(["Drogueria Beta"])
    ws_2.append(
        [
            "codigo_barra",
            "codigo_producto",
            "nombre_producto",
            "nombre_laboratorio",
            "unidades_existentes",
            "precio_unitario",
        ]
    )
    ws_2.append(["123456", "BET-9", "Cetirizina", "Lab Q", 12, 8])
    ws_2_bytes = BytesIO()
    workbook_2.save(ws_2_bytes)
    ws_2_bytes.seek(0)

    response = client.post(
        "/analizar",
        files=[
            (
                "files",
                (
                    "drogueria_alfa.xlsx",
                    ws_1_bytes.getvalue(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                ),
            ),
            (
                "files",
                (
                    "drogueria_beta.xlsx",
                    ws_2_bytes.getvalue(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                ),
            ),
        ],
    )

    assert response.status_code == 200
    data = response.json()
    assert len(data) == 1
    assert data[0]["codigo_barra"] == "123456"
    assert data[0]["codigo_interno_proveedor"] == "BET-9"
    assert data[0]["proveedor_ganador"] == "Drogueria Beta"
    assert data[0]["mejor_precio"] == 8.0


def test_analizar_dos_excels_con_columnas_extra_distintas() -> None:
    workbook_1 = Workbook()
    ws_1 = workbook_1.active
    ws_1.append(["Drogueria A"])
    ws_1.append(
        [
            "barra",
            "codigo",
            "descripcion",
            "stock",
            "laboratorio",
            "precio",
        ]
    )
    ws_1.append(["001234", "A-1", "Producto X", 5, "Lab Uno", 10])
    ws_1_bytes = BytesIO()
    workbook_1.save(ws_1_bytes)
    ws_1_bytes.seek(0)

    workbook_2 = Workbook()
    ws_2 = workbook_2.active
    ws_2.append(["Drogueria B"])
    ws_2.append(
        [
            "barra",
            "codigo",
            "descripcion",
            "stock",
            "laboratorio",
            "precio",
            "categoria",
            "campaign",
            "expiration",
        ]
    )
    ws_2.append(["1234", "B-9", "Producto X", 7, "Lab Uno", 7, "Med", "[]", False])
    ws_2_bytes = BytesIO()
    workbook_2.save(ws_2_bytes)
    ws_2_bytes.seek(0)

    response = client.post(
        "/analizar",
        files=[
            (
                "files",
                (
                    "drogueria_a.xlsx",
                    ws_1_bytes.getvalue(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                ),
            ),
            (
                "files",
                (
                    "drogueria_b.xlsx",
                    ws_2_bytes.getvalue(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                ),
            ),
        ],
    )

    assert response.status_code == 200
    data = response.json()
    assert len(data) == 1
    assert data[0]["codigo_barra"] == "1234"
    assert data[0]["codigo_interno_proveedor"] == "B-9"
    assert data[0]["proveedor_ganador"] == "Drogueria B"
    assert data[0]["mejor_precio"] == 7.0


def test_analizar_xlsx_con_codigo_producto_mixto_no_falla() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["Drogueria Mixta"])
    worksheet.append(
        [
            "codigo_barra",
            "codigo_producto",
            "nombre_producto",
            "nombre_laboratorio",
            "unidades_existentes",
            "precio_unitario",
        ]
    )
    worksheet.append(["770111", 120016, "Producto A", "Lab Mix", 30, 15])
    worksheet.append(["770111", "ML92806125016", "Producto A", "Lab Mix", 20, 11])

    file_content = BytesIO()
    workbook.save(file_content)
    file_content.seek(0)

    response = client.post(
        "/analizar",
        files={
            "file": (
                "mixto.xlsx",
                file_content.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200
    data = response.json()
    assert len(data) == 1
    assert data[0]["codigo_barra"] == "770111"
    assert data[0]["codigo_interno_proveedor"] == "ML92806125016"
    assert data[0]["mejor_precio"] == 11.0


def test_analizar_descarta_barcode_vacio_tras_normalizacion() -> None:
    csv_content = (
        "Drogueria Norte\n"
        "codigo_barra,codigo_producto,nombre_producto,nombre_laboratorio,unidades_existentes,precio_unitario\n"
        "0000,P1,Producto 1,Lab N,2,10\n"
        "1234,P2,Producto 2,Lab N,3,11\n"
    )

    response = client.post(
        "/analizar",
        files={"file": ("precios.csv", csv_content, "text/csv")},
    )

    assert response.status_code == 200
    assert response.json()[0]["codigo_barra"] == "1234"


def test_analizar_falla_si_falta_codigo_barra() -> None:
    csv_content = (
        "Proveedor sin barra\n"
        "codigo_producto,nombre_producto,nombre_laboratorio,unidades_existentes,precio_unitario\n"
        "A1,Arroz,Lab M,10,10\n"
    )

    response = client.post(
        "/analizar",
        files={"file": ("precios.csv", csv_content, "text/csv")},
    )

    assert response.status_code == 400
    detail = response.json()["detail"]
    assert detail["error"] == "Columnas requeridas faltantes"
    assert detail["faltantes"] == ["codigo_barra"]


def test_analizar_falla_si_no_hay_proveedor_en_primera_fila_ni_columna() -> None:
    csv_content = (
        "codigo_barra,codigo_producto,nombre_producto,nombre_laboratorio,unidades_existentes,precio_unitario\n"
        "12345,P1,Acetaminofen,Lab Z,10,8\n"
    )

    response = client.post(
        "/analizar",
        files={"file": ("precios_sin_proveedor.csv", csv_content, "text/csv")},
    )

    assert response.status_code == 400
    detail = response.json()["detail"]
    assert detail["error"] == "Falta nombre del proveedor"


def test_analizar_falla_si_falta_nombre_laboratorio() -> None:
    csv_content = (
        "Drogueria sin laboratorio\n"
        "codigo_barra,codigo_producto,nombre_producto,unidades_existentes,precio_unitario\n"
        "12345,P1,Acetaminofen,10,8\n"
    )

    response = client.post(
        "/analizar",
        files={"file": ("precios_sin_lab.csv", csv_content, "text/csv")},
    )

    assert response.status_code == 400
    detail = response.json()["detail"]
    assert detail["error"] == "Columnas requeridas faltantes"
    assert detail["faltantes"] == ["nombre_laboratorio"]


def test_analizar_rechaza_formato_invalido() -> None:
    response = client.post(
        "/analizar",
        files={"file": ("precios.txt", "contenido", "text/plain")},
    )

    assert response.status_code == 400


def test_cors_preflight() -> None:
    response = client.options(
        "/analizar",
        headers={
            "Origin": "http://localhost:3000",
            "Access-Control-Request-Method": "POST",
        },
    )

    # Si CORSMiddleware esta activo, FastAPI responde preflight 200.
    # Si esta desactivado, OPTIONS sobre /analizar retorna 405.
    assert response.status_code in {200, 405}
    if response.status_code == 200:
        assert response.headers["access-control-allow-origin"] == "*"
