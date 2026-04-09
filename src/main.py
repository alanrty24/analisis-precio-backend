import csv
import io
import os
from datetime import datetime, timezone
from pathlib import Path

import polars as pl
from fastapi import FastAPI, File, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from openpyxl import load_workbook

app = FastAPI(
	title="Price Optimizer API",
	description="Analizador masivo de mejores precios por proveedor",
	version="0.1.0",
)


def _parse_cors_origins() -> list[str]:
	raw_origins = os.getenv("CORS_ALLOW_ORIGINS", "*")
	return [origin.strip() for origin in raw_origins.split(",") if origin.strip()]


CORS_ALLOW_ORIGINS = _parse_cors_origins()
SUPPORTED_EXTENSIONS = {".csv", ".xlsx"}
MAX_UPLOAD_MB = int(os.getenv("MAX_UPLOAD_MB", "10"))
MAX_UPLOAD_BYTES = MAX_UPLOAD_MB * 1024 * 1024
REQUIRED_COLUMNS = {
	"codigo_barra",
	"codigo_producto",
	"nombre_producto",
	"nombre_laboratorio",
	"unidades_existentes",
	"precio_unitario",
}
BASE_PROCESSING_COLUMNS = [
	"codigo_barra",
	"codigo_producto",
	"nombre_producto",
	"nombre_laboratorio",
	"unidades_existentes",
	"precio_unitario",
	"nombre_proveedor",
]
COLUMN_ALIASES = {
	"barra": "codigo_barra",
	"barcode": "codigo_barra",
	"codigo_de_barra": "codigo_barra",
	"descripcion": "nombre_producto",
	"laboratorio": "nombre_laboratorio",
	"stock": "unidades_existentes",
	"existencia": "unidades_existentes",
	"unidades": "unidades_existentes",
	"disponible": "unidades_existentes",
	"precio": "precio_unitario",
	"codigo": "codigo_producto",
}

app.add_middleware(
	CORSMiddleware,
	allow_origins=CORS_ALLOW_ORIGINS,
	allow_credentials=False,
	allow_methods=["*"],
	allow_headers=["*"],
)


def _normalize_columns(df: pl.DataFrame) -> pl.DataFrame:
	renamed_columns = {
		column: COLUMN_ALIASES.get(column.strip().lower(), column.strip().lower())
		for column in df.columns
	}
	return df.rename(renamed_columns)


def _normalize_header_value(value: object) -> str:
	if value is None:
		return ""
	return str(value).strip().lower()


def _is_header_row(values: list[object]) -> bool:
	normalized_values = {
		COLUMN_ALIASES.get(_normalize_header_value(value), _normalize_header_value(value))
		for value in values
		if _normalize_header_value(value)
	}
	return (
		"codigo_barra" in normalized_values
		and "precio_unitario" in normalized_values
		and "codigo_producto" in normalized_values
	)


def _is_potential_header_row(values: list[object]) -> bool:
	normalized_values = {
		COLUMN_ALIASES.get(_normalize_header_value(value), _normalize_header_value(value))
		for value in values
		if _normalize_header_value(value)
	}
	header_candidates = set(REQUIRED_COLUMNS).union(COLUMN_ALIASES.values())
	return len(normalized_values.intersection(header_candidates)) >= 2


def _has_embedded_provider_row(first_row: list[object], second_row: list[object]) -> bool:
	non_empty_first_row = [value for value in first_row if _normalize_header_value(value)]
	return len(non_empty_first_row) == 1 and _is_potential_header_row(second_row)


def _build_dataframe_from_rows(headers: list[object], rows: list[tuple[object, ...]]) -> pl.DataFrame:
	normalized_headers = [
		_normalize_header_value(header) or f"column_{index + 1}"
		for index, header in enumerate(headers)
	]

	if not rows:
		return pl.DataFrame({header: [] for header in normalized_headers})

	# Forzamos texto en la lectura cruda para evitar conflictos de inferencia
	# (por ejemplo, codigo_producto numerico en algunas filas y alfanumerico en otras).
	column_count = len(normalized_headers)
	columns: dict[str, list[str | None]] = {header: [] for header in normalized_headers}
	for row in rows:
		for index in range(column_count):
			value = row[index] if index < len(row) else None
			columns[normalized_headers[index]].append(None if value is None else str(value))

	return pl.DataFrame(columns)


def _fill_provider_name(df: pl.DataFrame, provider_name: str | None) -> pl.DataFrame:
	if not provider_name:
		return df

	if "nombre_proveedor" not in df.columns:
		return df.with_columns(pl.lit(provider_name).alias("nombre_proveedor"))

	provider_column = pl.col("nombre_proveedor").cast(pl.Utf8, strict=False).str.strip_chars()
	return df.with_columns(
		pl.when(provider_column.is_null() | (provider_column == ""))
		.then(pl.lit(provider_name))
		.otherwise(provider_column)
		.alias("nombre_proveedor")
	)


def _ensure_provider_column(df: pl.DataFrame, provider_name: str | None) -> pl.DataFrame:
	df_with_provider = _fill_provider_name(df, provider_name)
	if "nombre_proveedor" in df_with_provider.columns:
		return df_with_provider

	raise HTTPException(
		status_code=400,
		detail={
			"error": "Falta nombre del proveedor",
			"detalle": "Coloque el nombre del proveedor en la fila 1, columna A, o incluya la columna nombre_proveedor.",
		},
	)


def _validate_required_columns(df: pl.DataFrame) -> None:
	missing_columns = sorted(REQUIRED_COLUMNS.difference(df.columns))
	if missing_columns:
		raise HTTPException(
			status_code=400,
			detail={
				"error": "Columnas requeridas faltantes",
				"faltantes": missing_columns,
			},
		)


def _read_csv_input(content: bytes) -> tuple[pl.DataFrame, str | None]:
	text_stream = io.StringIO(content.decode("utf-8-sig"))
	reader = csv.reader(text_stream)
	first_row = next(reader, [])
	second_row = next(reader, [])

	if _has_embedded_provider_row(first_row, second_row):
		provider_name = str(first_row[0]).strip()
		return pl.read_csv(io.BytesIO(content), skip_rows=1), provider_name

	return pl.read_csv(io.BytesIO(content)), None


def _read_excel_input(content: bytes) -> tuple[pl.DataFrame, str | None]:
	workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
	worksheet = workbook.active
	rows = worksheet.iter_rows(values_only=True)
	first_row = list(next(rows, ()))
	second_row = list(next(rows, ()))

	if _has_embedded_provider_row(first_row, second_row):
		provider_name = str(first_row[0]).strip()
		data_rows = list(rows)
		workbook.close()
		return _build_dataframe_from_rows(second_row, data_rows), provider_name

	workbook.close()
	return pl.read_excel(io.BytesIO(content), engine="openpyxl"), None


def _read_input_file(filename: str, content: bytes) -> tuple[pl.DataFrame, str | None]:
	extension = Path(filename).suffix.lower()

	if extension == ".csv":
		return _read_csv_input(content)

	return _read_excel_input(content)


def _normalize_barcode_expression(column_name: str) -> pl.Expr:
	return (
		pl.col(column_name)
		.cast(pl.Utf8, strict=False)
		.str.replace_all(r"\s+", "")
		.str.replace_all(r"^0+", "")
	)


@app.get("/")
async def health_check() -> dict[str, str]:
	return {"status": "ready", "engine": "polars-rust"}


def _validate_input_file(upload_file: UploadFile) -> None:
	if not upload_file.filename:
		raise HTTPException(status_code=400, detail="El archivo debe tener nombre")

	extension = Path(upload_file.filename).suffix.lower()
	if extension not in SUPPORTED_EXTENSIONS:
		raise HTTPException(
			status_code=400,
			detail="Formato no soportado. Use archivos .csv o .xlsx",
		)


@app.post("/analizar")
async def analizar_precios(
	file: UploadFile | None = File(default=None),
	files: list[UploadFile] | None = File(default=None),
) -> list[dict[str, object]]:
	input_files: list[UploadFile] = []
	if file is not None:
		input_files.append(file)
	if files:
		input_files.extend(files)

	if not input_files:
		raise HTTPException(
			status_code=400,
			detail="Debe enviar al menos un archivo en 'file' o 'files'.",
		)

	try:
		dfs: list[pl.DataFrame] = []
		for current_file in input_files:
			_validate_input_file(current_file)

			content = await current_file.read()
			if not content:
				raise HTTPException(
					status_code=400,
					detail=f"El archivo '{current_file.filename}' esta vacio",
				)

			if len(content) > MAX_UPLOAD_BYTES:
				raise HTTPException(
					status_code=413,
					detail=f"El archivo '{current_file.filename}' excede el limite de {MAX_UPLOAD_MB} MB",
				)

			df, provider_name = _read_input_file(current_file.filename, content)
			df = _normalize_columns(df)
			df = _ensure_provider_column(df, provider_name)
			_validate_required_columns(df)
			df = df.select(BASE_PROCESSING_COLUMNS)
			dfs.append(df)

		df = pl.concat(dfs, how="vertical_relaxed")
		analysis_timestamp = datetime.now(timezone.utc).isoformat()

		resultado = (
			df.with_columns(
				_normalize_barcode_expression("codigo_barra").alias("codigo_barra"),
				pl.col("precio_unitario").cast(pl.Float64, strict=False),
				pl.col("codigo_producto").cast(pl.Utf8, strict=False).str.strip_chars(),
				pl.col("nombre_producto").cast(pl.Utf8, strict=False).str.strip_chars(),
				pl.col("nombre_laboratorio").cast(pl.Utf8, strict=False).str.strip_chars(),
				pl.col("nombre_proveedor").cast(pl.Utf8, strict=False).str.strip_chars(),
				pl.col("unidades_existentes").cast(pl.Int64, strict=False),
			)
			.filter(
				pl.col("codigo_barra").is_not_null()
				& (pl.col("codigo_barra") != "")
				& pl.col("precio_unitario").is_not_null()
				& (pl.col("precio_unitario") > 0)
			)
			.sort(["codigo_barra", "precio_unitario", "nombre_proveedor", "codigo_producto"])
			.unique(subset=["codigo_barra"], keep="first", maintain_order=True)
			.select(
				pl.col("codigo_barra"),
				pl.col("codigo_producto").alias("codigo_interno_proveedor"),
				pl.col("nombre_producto"),
				pl.col("precio_unitario").alias("mejor_precio"),
				pl.col("nombre_proveedor").alias("proveedor_ganador"),
				pl.col("unidades_existentes").alias("unidades_disponibles"),
			)
			.with_columns(pl.lit(analysis_timestamp).alias("analisis_timestamp"))
		)

		return resultado.to_dicts()
	except HTTPException:
		raise
	except pl.exceptions.PolarsError as exc:
		raise HTTPException(
			status_code=400,
			detail={
				"error": "No se pudo interpretar el archivo de entrada",
				"detalle": str(exc),
			},
		) from exc
	except Exception as exc:
		raise HTTPException(
			status_code=500,
			detail={
				"error": "Fallo en el procesamiento",
				"detalle": str(exc),
			},
		) from exc
